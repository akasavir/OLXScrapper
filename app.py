import requests
from bs4 import BeautifulSoup
import re
import random
import time
from datetime import datetime
import pandas as pd
import os
import subprocess
import platform
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Pobieranie miast od użytkownika
cities_input = input("Podaj miasta do sprawdzenia, oddzielone przecinkami (bez polskich znaków!): ")
cities = [city.strip().lower() for city in cities_input.split(',')]  # Tworzymy listę miast i usuwamy nadmiarowe spacje

# Pobieranie maksymalnej ceny od użytkownika
max_price_input = input("Podaj maksymalną cenę (w zł): ")
try:
    max_price = float(max_price_input.replace(" ", ""))  # Usuwamy spacje z ceny, jeśli są
except ValueError:
    print("Podano nieprawidłową cenę. Program zakończy działanie.")
    exit()

# Nagłówki, aby imitować zapytanie z przeglądarki
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
}

# Funkcja do losowego opóźnienia
def random_delay():
    delay = random.randint(3, 10)  # Opóźnienie między 3 a 10 sekund
    print(f"Opóźnienie: {delay} sekund...")
    time.sleep(delay)  # Opóźnienie

# Zmienna do przechowywania wyników
results = []

# Iterujemy po miastach
for city in cities:
    print(f"Przetwarzam miasto: {city}...")
    url = f"https://www.olx.pl/nieruchomosci/mieszkania/sprzedaz/{city}/"
    
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        print(f"Błąd podczas pobierania strony dla {city}: {response.status_code}")
        continue

    print(f"Strona {city} została pobrana pomyślnie.")

    soup = BeautifulSoup(response.text, 'html.parser')
    
    # Szukamy ogłoszeń
    ads = soup.find_all('div', {'data-cy': 'l-card'})  # Poprawne selektory dla OLX
    
    if not ads:
        print(f"Brak wyników dla {city}.")
        continue

    # Iteracja przez ogłoszenia
    for ad in ads:
        # Cena
        price_tag = ad.find('p', {'data-testid': 'ad-price'})
        price = price_tag.get_text(strip=True) if price_tag else "Brak ceny"

        # Lokalizacja
        location_tag = ad.find('p', {'data-testid': 'location-date'})
        location = location_tag.get_text(strip=True) if location_tag else "Brak lokalizacji"

        # Link
        link_tag = ad.find('a', {'class': 'css-qo0cxu'})
        link = link_tag['href'] if link_tag else "Brak linku"
        if not link.startswith("http"):
            link = "https://www.olx.pl" + link

        # Usunięcie frazy "do negocjacji" z ceny
        price = price.replace(" do negocjacji", "").strip()

        # Wyciągnięcie liczby z ceny
        match = re.search(r'(\d+(\s?\d{3})*)', price)
        if match:
            price_value = float(match.group(1).replace(" ", ""))

            # Dodanie ogłoszenia do wyników, jeśli cena mieści się w limicie
            if price_value <= max_price:
                results.append({
                    "Miejscowość": location,
                    "Cena": price_value,
                    "Link": link
                })

    # Dodajemy opóźnienie po przetworzeniu strony
    random_delay()

# Jeśli znaleziono wyniki, zapisujemy je do pliku Excel
if results:
    # Tworzymy unikalną nazwę pliku na podstawie daty i godziny
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f"[Mieszkania na sprzedaż]{timestamp}.xlsx"
    
    # Zapisujemy wyniki do pliku Excel
    df = pd.DataFrame(results)
    df.to_excel(file_name, index=False)

    # Ładowanie pliku Excel do openpyxl
    wb = load_workbook(file_name)
    ws = wb.active
    
    # Określamy zakres komórek w kolumnie "Cena" (zakładamy, że ceny są w kolumnie B)
    col_range = f"B2:B{len(results)+1}"
    
    # Tworzymy wypełnienia (tła) dla formatowania warunkowego
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Zielony
    red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Czerwony
    
    # Dodajemy formatowanie warunkowe
    for row in ws[col_range]:
        for cell in row:
            if cell.value <= max_price:
                cell.fill = green_fill  # Zielone tło dla cen <= max_price
            else:
                cell.fill = red_fill  # Czerwone tło dla cen > max_price
    
    # Zapisujemy plik z formatowaniem
    wb.save(file_name)
    
    print(f"Wyniki zapisane do pliku {file_name} z formatowaniem.")

    # Otwórz folder lub plik
    if platform.system() == "Windows":
        subprocess.run(["explorer", "/select,", os.path.abspath(file_name)])
    elif platform.system() == "Darwin":  # macOS
        subprocess.run(["open", os.path.abspath(file_name)])
    else:  # Linux
        subprocess.run(["xdg-open", os.path.abspath(file_name)])
else:
    print("Nie znaleziono żadnych ofert pasujących do wzorca.")
